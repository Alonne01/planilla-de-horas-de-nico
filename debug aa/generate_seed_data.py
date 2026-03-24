#!/usr/bin/env python3
"""
Cross-reference Excel sector data with PDF personnel data.
Generate TypeScript EMPLEADOS array for seed.ts.
"""
import json, re, unicodedata, sys
from pathlib import Path
import openpyxl

BASE = Path(r"C:\Users\alonn\OneDrive\Documents\WENLEN\programas\planilla de horas\debug aa")
OUT_TS = BASE / "empleados_output.ts"
OUT_CATS = BASE / "categorias_output.ts"


# ═══════════════════════════════════════════════════════
# UTILITIES
# ═══════════════════════════════════════════════════════

def norm(s):
    """Remove accents, lowercase, collapse spaces."""
    if not s:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()


def parse_apellido_nombre(raw):
    """Parse 'APELLIDO, NOMBRE' into (apellido, nombre)."""
    raw = raw.strip()
    if "," in raw:
        parts = raw.split(",", 1)
        return parts[0].strip(), parts[1].strip()
    else:
        # No comma — take first word as apellido, rest as nombre
        parts = raw.split(None, 1)
        if len(parts) == 2:
            return parts[0].strip(), parts[1].strip()
        return raw, ""


def titlecase_name(s):
    """Proper titlecase for Argentine names."""
    if not s:
        return ""
    particles = {"de", "del", "la", "las", "los", "y", "e"}
    words = s.split()
    result = []
    for i, w in enumerate(words):
        wl = w.lower()
        if wl in particles and i > 0:
            result.append(wl)
        elif wl.startswith("mc") and len(wl) > 2:
            result.append("Mc" + w[2:].capitalize())
        else:
            result.append(w.capitalize())
    return " ".join(result)


def make_email(nombre, apellido, used_emails):
    """Generate unique email."""
    # Clean first name → use first word
    fn = norm(nombre).split()[0] if nombre else "x"
    # Clean apellido → use first word (handle compound)
    ap = norm(apellido).split()[0] if apellido else "x"
    # Remove non-alpha
    fn = re.sub(r"[^a-z]", "", fn)
    ap = re.sub(r"[^a-z]", "", ap)
    
    base = f"{fn}.{ap}"
    email = f"{base}@wenlen.com"
    
    if email in used_emails:
        # Add second name initial or number
        nombre_parts = norm(nombre).split()
        if len(nombre_parts) > 1:
            fn2 = re.sub(r"[^a-z]", "", nombre_parts[1])
            email = f"{fn}.{fn2[0]}.{ap}@wenlen.com" if fn2 else f"{fn}.{ap}2@wenlen.com"
        if email in used_emails:
            i = 2
            while f"{base}{i}@wenlen.com" in used_emails:
                i += 1
            email = f"{base}{i}@wenlen.com"
    
    used_emails.add(email)
    return email


# ═══════════════════════════════════════════════════════
# 1. LOAD PDF DATA
# ═══════════════════════════════════════════════════════

print("Loading PDF data...")
pdf_data = json.load(open(BASE / "pdf_people.json", "r", encoding="utf-8"))
print(f"  {len(pdf_data)} people in PDF")

# Build lookup: normalized "apellido nombre" → pdf entry
pdf_by_norm = {}
for p in pdf_data:
    ap, nom = parse_apellido_nombre(p["nombre"])
    key = norm(f"{ap} {nom}")
    pdf_by_norm[key] = {**p, "_ap": ap, "_nom": nom}

# Also build lookup by just normalized apellido (for fuzzy matching)
pdf_by_apellido = {}
for p in pdf_data:
    ap, nom = parse_apellido_nombre(p["nombre"])
    key = norm(ap)
    if key not in pdf_by_apellido:
        pdf_by_apellido[key] = []
    pdf_by_apellido[key].append({**p, "_ap": ap, "_nom": nom})


# ═══════════════════════════════════════════════════════
# 2. PARSE EXCEL
# ═══════════════════════════════════════════════════════

print("\nParsing Excel...")
wb = openpyxl.load_workbook(BASE / "Personal - Sectores.xlsx", data_only=True)
ws = wb.active

# Sector mapping: Excel sector key → App sector name
SECTOR_MAP = {
    "WIRELINE": "Wireline",
    "CMASS": "CMASS",
    "INTENDENCIA": "Intendencia",
    "MANTENIMIENTO": "Logística y Transporte",  # No own supervisors → merge
    "CAMIONEROS": "Logística y Transporte",      # No own supervisors → merge
    "LOGISTICA": "Logística y Transporte",
    "ADMINISTRACION": "Administración",
    "ALMACEN": "Almacén",
    "TESTING": "Testing",
    "CABEZALES": "Cabezales",
    "FRACTURA": "Fractura",
}

# Sub-headers: (row_1indexed, col_0indexed) → new_role
SUB_HEADERS = {
    "Supervisores": True,
    "Supervisores FRACTURA": True,
    "FRACTURA": True,
    "Supervisores y Gerencia Testing": True,
    "Supervisores y Gerencia cabezales": True,
    "Coordinadores FRACTURA": True,
    "Ord": True,
}

# Known role overrides by (normalized_apellido, sector_key) or (norm_full_name)
ROLE_OVERRIDES = {}

# CMASS supervisors
for n in ["ATIENZA", "MERIÑO"]:
    ROLE_OVERRIDES[(norm(n), "CMASS")] = "SUPERVISOR"

# Logística coordinator
ROLE_OVERRIDES[(norm("BELTRAME"), "LOGISTICA", norm("MAURICIO"))] = "COORDINADOR"

# Testing coordinators
for n in ["ARISMENDI", "CORACHAN"]:
    ROLE_OVERRIDES[(norm(n), "TESTING")] = "COORDINADOR"

# Cabezales coordinators (Ávila and Kloberdanz Abelardo)
ROLE_OVERRIDES[(norm("ÁVILA"), "CABEZALES")] = "COORDINADOR"
ROLE_OVERRIDES[(norm("KLOBERDANZ"), "CABEZALES", norm("ABELARDO"))] = "COORDINADOR"

# Admin RRHH
for n in ["AGÜERO", "CEJAS", "STRILLEVSKY"]:
    ROLE_OVERRIDES[(norm(n), "ADMINISTRACION")] = "RRHH"
ROLE_OVERRIDES[(norm("GONZÁLEZ"), "ADMINISTRACION", norm("AMALIA"))] = "RRHH"

# Admin Gerentes
for n in ["WINKLER", "SILVEIRA"]:
    ROLE_OVERRIDES[(norm(n), "ADMINISTRACION")] = "GERENTE"
ROLE_OVERRIDES[(norm("DÍAZ"), "ADMINISTRACION", norm("CARLOS"))] = "GERENTE"


def get_override_role(apellido_raw, nombre_raw, sector_key):
    """Check if there's a role override for this person."""
    ap_norm = norm(apellido_raw)
    nom_norm = norm(nombre_raw)
    
    # Check 3-key overrides first (apellido + sector + partial nombre)
    for key, role in ROLE_OVERRIDES.items():
        if len(key) == 3:
            k_ap, k_sec, k_nom_partial = key
            if k_ap == ap_norm and k_sec == sector_key and k_nom_partial in nom_norm:
                return role
    
    # Check 2-key overrides
    key2 = (ap_norm, sector_key)
    if key2 in ROLE_OVERRIDES:
        return ROLE_OVERRIDES[key2]
    
    return None


# Parse each column
# Column definitions: (ord_col_0idx, name_col_0idx, sector_key, initial_role, role_transitions)
# role_transitions: {row: new_role}
COLUMNS = [
    {"oc": 0, "nc": 1, "sector": "WIRELINE", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 2, "nc": 3, "sector": "CMASS", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 4, "nc": 5, "sector": "INTENDENCIA", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 6, "nc": 7, "sector": "MANTENIMIENTO", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 8, "nc": 9, "sector": "CAMIONEROS", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 10, "nc": 11, "sector": "LOGISTICA", "init_role": "SUPERVISOR", "transitions": {}},
    {"oc": 12, "nc": 13, "sector": "LOGISTICA", "init_role": "OPERADOR", "transitions": {}},  # Asistentes
    {"oc": 14, "nc": 15, "sector": "ADMINISTRACION", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 16, "nc": 17, "sector": "ALMACEN", "init_role": "OPERADOR", "transitions": {}},
    {"oc": 18, "nc": 19, "sector": "TESTING", "init_role": "OPERADOR", "transitions": {50: "SUPERVISOR"}},
    {"oc": 20, "nc": 21, "sector": "CABEZALES", "init_role": "OPERADOR", "transitions": {49: "SUPERVISOR"}},
    {"oc": 22, "nc": 23, "sector": "FRACTURA", "init_role": "OPERADOR", "transitions": {}},  # Oper. PH
    {"oc": 24, "nc": 25, "sector": "FRACTURA", "init_role": "COORDINADOR", "transitions": {6: "SUPERVISOR", 25: "OPERADOR"}},
]

excel_people = []

for col_def in COLUMNS:
    nc = col_def["nc"]
    sector_key = col_def["sector"]
    current_role = col_def["init_role"]
    transitions = col_def["transitions"]
    
    for r in range(2, ws.max_row + 1):
        # Check for role transition at this row
        if r in transitions:
            current_role = transitions[r]
        
        name_val = ws.cell(row=r, column=nc + 1).value  # openpyxl uses 1-indexed
        if not name_val or not isinstance(name_val, str):
            continue
        name_val = name_val.strip()
        
        # Skip sub-headers and labels
        if name_val in SUB_HEADERS or len(name_val) < 4:
            continue
        
        # Parse name
        apellido_raw, nombre_raw = parse_apellido_nombre(name_val)
        
        # Skip if it's clearly not a person name
        if not nombre_raw and apellido_raw.isupper() and " " not in apellido_raw:
            continue
        
        # Apply role override
        override = get_override_role(apellido_raw, nombre_raw, sector_key)
        role = override if override else current_role
        
        excel_people.append({
            "apellido_raw": apellido_raw,
            "nombre_raw": nombre_raw,
            "full_raw": name_val,
            "sector_key": sector_key,
            "role": role,
            "row": r,
            "col": nc,
        })

print(f"  {len(excel_people)} people in Excel")

# Count by sector
from collections import Counter
sector_counts = Counter(p["sector_key"] for p in excel_people)
for s, c in sorted(sector_counts.items()):
    print(f"    {s}: {c}")


# ═══════════════════════════════════════════════════════
# 3. CROSS-REFERENCE EXCEL ↔ PDF
# ═══════════════════════════════════════════════════════

print("\nCross-referencing...")
matched = []
unmatched_excel = []
used_pdf_keys = set()

for ep in excel_people:
    ap_norm = norm(ep["apellido_raw"])
    nom_norm = norm(ep["nombre_raw"])
    full_key = norm(f"{ep['apellido_raw']} {ep['nombre_raw']}")
    
    # Try exact full match
    pdf_match = pdf_by_norm.get(full_key)
    
    if not pdf_match:
        # Try with just apellido + first nombre word
        nom_first = nom_norm.split()[0] if nom_norm else ""
        partial_key = f"{ap_norm} {nom_first}"
        pdf_match = pdf_by_norm.get(partial_key)
    
    if not pdf_match:
        # Try partial apellido match (e.g., "PARRA" matching "PARRA MONSALVE")
        for pkey, pval in pdf_by_norm.items():
            if pkey.startswith(ap_norm + " ") and nom_norm and nom_norm.split()[0] in pkey:
                pdf_match = pval
                break
            # Also try if PDF apellido starts with Excel apellido
            pdf_ap = norm(pval["_ap"])
            if pdf_ap.startswith(ap_norm) and len(pdf_ap) > len(ap_norm):
                pdf_nom = norm(pval["_nom"])
                if nom_norm and nom_norm.split()[0] in pdf_nom:
                    pdf_match = pval
                    break
    
    if not pdf_match:
        # Try apellido-only match (if unique)
        candidates = pdf_by_apellido.get(ap_norm, [])
        if len(candidates) == 1:
            pdf_match = candidates[0]
        elif len(candidates) > 1:
            # Try to disambiguate by first nombre
            nom_first = nom_norm.split()[0] if nom_norm else ""
            for c in candidates:
                c_nom = norm(c["_nom"]).split()[0] if c["_nom"] else ""
                if c_nom == nom_first:
                    pdf_match = c
                    break
            if not pdf_match:
                # Try partial nombre match
                for c in candidates:
                    c_nom = norm(c["_nom"])
                    if nom_first and nom_first in c_nom:
                        pdf_match = c
                        break
    
    if pdf_match:
        match_key = norm(f"{pdf_match['_ap']} {pdf_match['_nom']}")
        used_pdf_keys.add(match_key)
        matched.append({
            **ep,
            "legajo": pdf_match["legajo"],
            "dni": pdf_match["dni"],
            "telefono": pdf_match["telefono"],
            "fecha_ingreso": pdf_match["alta"],
            "categoria_pdf": pdf_match["categoria"],
            "nombre_pdf": pdf_match["_nom"],
            "apellido_pdf": pdf_match["_ap"],
        })
    else:
        unmatched_excel.append(ep)

# Find PDF people not in Excel
unmatched_pdf = []
for key, p in pdf_by_norm.items():
    if key not in used_pdf_keys:
        unmatched_pdf.append(p)

print(f"  Matched: {len(matched)}")
print(f"  Unmatched Excel: {len(unmatched_excel)}")
if unmatched_excel:
    for u in unmatched_excel:
        print(f"    ❌ Excel: {u['full_raw']} (sector={u['sector_key']}, role={u['role']})")
print(f"  Unmatched PDF: {len(unmatched_pdf)}")
if unmatched_pdf:
    for u in unmatched_pdf[:20]:
        print(f"    ❌ PDF: {u['nombre']} (legajo={u['legajo']}, cat={u['categoria']})")
    if len(unmatched_pdf) > 20:
        print(f"    ... and {len(unmatched_pdf) - 20} more")


# ═══════════════════════════════════════════════════════
# 4. CATEGORY MAPPING
# ═══════════════════════════════════════════════════════

# Map PDF categories to seed category codes
# PDF format: "TII TA Cat VII", "TII TB Cat III", "TIII TS Cat VIII", "SPJ"
# Seed format: For PP (CCT 644/12) we'll create matching codes
# For PJ (CCT 637/11): SPJ → JER-B (default)

def map_categoria(cat_pdf):
    """Map PDF category string to (convenio, cat_code, cat_nombre)."""
    if not cat_pdf:
        return ("PP", "TIII-C", "Título III — Operador 2do / Técnico")
    
    cat_pdf = cat_pdf.strip()
    
    if cat_pdf == "SPJ":
        return ("PJ", "SPJ", "Sector Petrolero Jerárquico")
    
    # Parse "TII TA Cat VII" → titulo=TII, tipo=TA, catnum=VII
    m = re.match(r"(T\w+)\s+(T[AB]|TS)\s+Cat\s+(\w+)", cat_pdf)
    if m:
        titulo, tipo, num = m.groups()
        code = f"{titulo}-{tipo}-{num}"
        
        roman_to_int = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
                        "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11}
        orden_num = roman_to_int.get(num, 0)
        
        if titulo == "TII" and tipo == "TA":
            nombre = f"Título II Tipo A — Categoría {num}"
            orden = orden_num
        elif titulo == "TII" and tipo == "TB":
            nombre = f"Título II Tipo B — Categoría {num}"
            orden = 20 + orden_num
        elif titulo == "TIII" and tipo == "TS":
            nombre = f"Título III Téc. y Serv. — Categoría {num}"
            orden = 40 + orden_num
        else:
            nombre = cat_pdf
            orden = 50 + orden_num
        
        return ("PP", code, nombre, orden)
    
    # Fallback
    return ("PP", "TIII-C", "Título III — Operador 2do / Técnico", 12)


# Collect all unique categories
all_cats = set()
for m in matched:
    all_cats.add(m["categoria_pdf"])
for u in unmatched_pdf:
    all_cats.add(u["categoria"])

print(f"\n  Unique categories: {len(all_cats)}")


# ═══════════════════════════════════════════════════════
# 5. GENERATE TYPESCRIPT OUTPUT
# ═══════════════════════════════════════════════════════

print("\nGenerating TypeScript...")

# Combine matched + unmatched_pdf (assign to FRACTURA OPERADOR by default for unmatched)
all_employees = []
used_emails = set()

# Sort matched by sector then apellido
matched.sort(key=lambda x: (x["sector_key"], norm(x.get("apellido_pdf", x["apellido_raw"]))))

for m in matched:
    apellido = titlecase_name(m.get("apellido_pdf", m["apellido_raw"]))
    nombre = titlecase_name(m.get("nombre_pdf", m["nombre_raw"]))
    sector_key = m["sector_key"]
    app_sector = SECTOR_MAP[sector_key]
    role = m["role"]
    
    # For RRHH/GERENTE in Administración, sector stays ADMINISTRACION but sectorId will be null
    
    email = make_email(nombre, apellido, used_emails)
    
    cat_info = map_categoria(m["categoria_pdf"])
    convenio = cat_info[0]  # "PP" or "PJ"
    cat_code = cat_info[1]
    
    # RRHH/GERENTE/COORDINADOR/SUPERVISOR are jerárquicos → use PJ if category is SPJ
    jerarquico_roles = ["RRHH", "GERENTE", "COORDINADOR", "SUPERVISOR"]
    if role in jerarquico_roles and convenio != "PJ":
        # They might still be PP — keep their actual category
        pass
    
    # Parse fecha_ingreso
    fecha = m.get("fecha_ingreso", "01/01/2024")
    if fecha:
        # Convert DD/MM/YYYY to YYYY-MM-DD
        parts = fecha.split("/")
        if len(parts) == 3:
            fecha_iso = f"{parts[2]}-{parts[1]}-{parts[0]}"
        else:
            fecha_iso = "2024-01-01"
    else:
        fecha_iso = "2024-01-01"
    
    all_employees.append({
        "nombre": nombre,
        "apellido": apellido,
        "sector": sector_key,
        "rol": role,
        "legajo": m["legajo"],
        "email": email,
        "dni": m.get("dni", ""),
        "telefono": m.get("telefono", ""),
        "fechaIngreso": fecha_iso,
        "categoria": cat_code,
        "convenio": convenio,
    })

# Add unmatched PDF people (not in Excel — might have left the company or been missed)
for u in unmatched_pdf:
    apellido = titlecase_name(u["_ap"])
    nombre = titlecase_name(u["_nom"])
    email = make_email(nombre, apellido, used_emails)
    
    cat_info = map_categoria(u["categoria"])
    convenio = cat_info[0]
    cat_code = cat_info[1]
    
    fecha = u.get("alta", "01/01/2024")
    if fecha:
        parts = fecha.split("/")
        if len(parts) == 3:
            fecha_iso = f"{parts[2]}-{parts[1]}-{parts[0]}"
        else:
            fecha_iso = "2024-01-01"
    else:
        fecha_iso = "2024-01-01"
    
    all_employees.append({
        "nombre": nombre,
        "apellido": apellido,
        "sector": "FRACTURA",  # Default sector for unmatched
        "rol": "OPERADOR",
        "legajo": u["legajo"],
        "email": email,
        "dni": u.get("dni", ""),
        "telefono": u.get("telefono", ""),
        "fechaIngreso": fecha_iso,
        "categoria": cat_code,
        "convenio": convenio,
        "_unmatched": True,
    })


# ═══════════════════════════════════════════════════════
# 6. WRITE OUTPUT
# ═══════════════════════════════════════════════════════

# Write EMPLEADOS TypeScript
lines = []
lines.append("// ═══════════════════════════════════════════════════════════════")
lines.append("// EMPLEADOS — Nómina completa generada desde Excel + PDF")
lines.append("// ═══════════════════════════════════════════════════════════════")
lines.append("const EMPLEADOS: {")
lines.append("  nombre: string; apellido: string; sector: string; rol: string;")
lines.append("  legajo: string; email: string; dni: string; telefono: string;")
lines.append("  fechaIngreso: string; categoria: string; convenio: string;")
lines.append("}[] = [")

current_sector = None
for emp in all_employees:
    if emp.get("_unmatched"):
        continue  # Skip unmatched PDF-only people for now
    
    sector = emp["sector"]
    if sector != current_sector:
        lines.append(f"  // -- {sector} --")
        current_sector = sector
    
    # Escape single quotes
    nombre = emp["nombre"].replace("'", "\\'")
    apellido = emp["apellido"].replace("'", "\\'")
    email = emp["email"]
    legajo = emp["legajo"]
    rol = emp["rol"]
    dni = emp["dni"].replace("\u2010", "-").replace("\u2011", "-").replace("\u2012", "-").replace("\u2013", "-")
    telefono = emp["telefono"].replace("\u2010", "-").replace("\u2011", "-").replace("\u2012", "-").replace("\u2013", "-")
    fecha = emp["fechaIngreso"]
    cat = emp["categoria"]
    conv = emp["convenio"]
    
    line = (f"  {{ nombre: '{nombre}', apellido: '{apellido}', "
            f"sector: '{sector}', rol: '{rol}', "
            f"legajo: '{legajo}', email: '{email}', "
            f"dni: '{dni}', telefono: '{telefono}', "
            f"fechaIngreso: '{fecha}', categoria: '{cat}', convenio: '{conv}' }},")
    lines.append(line)

lines.append("];")

with open(OUT_TS, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"\n✅ Generated {OUT_TS}")
print(f"   {len([e for e in all_employees if not e.get('_unmatched')])} matched employees")
print(f"   {len([e for e in all_employees if e.get('_unmatched')])} unmatched (PDF-only, excluded)")

# Write category definitions
cat_defs = {}
for emp in all_employees:
    if emp.get("_unmatched"):
        continue
    cat_pdf = None
    for m in matched:
        if m["legajo"] == emp["legajo"]:
            cat_pdf = m["categoria_pdf"]
            break
    if cat_pdf:
        info = map_categoria(cat_pdf)
        if len(info) == 4:
            _, code, nombre, orden = info
        else:
            _, code, nombre = info
            orden = 0
        if code not in cat_defs:
            cat_defs[code] = {"codigo": code, "nombre": nombre, "orden": orden, "convenio": info[0]}

# Sort and write
pp_cats = sorted([c for c in cat_defs.values() if c["convenio"] == "PP"], key=lambda x: x["orden"])
pj_cats = sorted([c for c in cat_defs.values() if c["convenio"] == "PJ"], key=lambda x: x["orden"])

with open(OUT_CATS, "w", encoding="utf-8") as f:
    f.write("// CCT 644/12 Categories (from PDF data)\n")
    f.write("const catsPP = [\n")
    for c in pp_cats:
        f.write(f"  {{ codigo: '{c['codigo']}', nombre: '{c['nombre']}', orden: {c['orden']} }},\n")
    f.write("];\n\n")
    f.write("// CCT 637/11 Categories (from PDF data)\n")
    f.write("const catsPJ = [\n")
    for c in pj_cats:
        f.write(f"  {{ codigo: '{c['codigo']}', nombre: '{c['nombre']}', orden: {c['orden']} }},\n")
    f.write("];\n")

print(f"✅ Generated {OUT_CATS}")
print(f"   PP categories: {len(pp_cats)}")
print(f"   PJ categories: {len(pj_cats)}")

# Summary
print("\n═══════════════════════════════════════")
print("SUMMARY")
print("═══════════════════════════════════════")
role_counts = Counter(e["rol"] for e in all_employees if not e.get("_unmatched"))
for r, c in sorted(role_counts.items()):
    print(f"  {r}: {c}")
sector_role = Counter((e["sector"], e["rol"]) for e in all_employees if not e.get("_unmatched"))
print("\nBy sector × role:")
for (s, r), c in sorted(sector_role.items()):
    print(f"  {s} / {r}: {c}")
